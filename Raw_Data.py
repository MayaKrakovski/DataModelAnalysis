import openpyxl
import os
import csv
import pandas as pd
import matplotlib.pyplot as plt
import math
import numpy as np
from scipy.signal import find_peaks
from scipy.signal import butter, filtfilt


def min_max_scale_signal(signal, exercise_range, desired_min=0, desired_max=1, by_dict=False):
    # by_dict - True if the scaling is by the predefined values of each exercise
    # False - if it is by the min/max value of the signal
    if by_dict:
        min_val, max_val = exercise_range[0], exercise_range[1]
    else:
        min_val = signal.min()
        max_val = signal.max()
    scaled_signal = (signal - min_val) / (max_val - min_val) * (desired_max - desired_min) + desired_min
    return scaled_signal


def scale_raw_data():
    exercises = {'raise_arms_horizontally': [30, 120], 'bend_elbows': [0, 180],'raise_arms_bend_elbows': [0, 180]}
    data_sources = ['maya', 'naama', 'naama_pilot', 'val1']

    df = pd.DataFrame()
    for d in data_sources:
        for e in exercises:
            temp = pd.read_csv('CSV/Raw Data/'+d+'_'+e+'.csv')
            signalcols = temp.columns[temp.columns.str.startswith('Unnamed')]
            signal = temp[signalcols]
            scaled_signal = min_max_scale_signal(signal, exercises[e], 0, 1, True)
            temp[signalcols] = scaled_signal
            temp.insert(1, "Exercise", [e]*len(temp.index), True)
            temp.insert(1, "Source", [d]*len(temp.index), True)
            df = df.append(temp, ignore_index=True)

    df.to_csv('CSV\\Raw Data\\all_raw_data_scaled_by_local_minmax.csv')

    # plotting
    framepersec = 30  # frame per seconds of nuitrack
    # For filtering
    cutoff_freq = 6
    filter_order = 2
    y, x = butter(filter_order, cutoff_freq/(framepersec/2))
    plt.style.use('ggplot')
    df = pd.read_csv('CSV/Raw Data/all_raw_data_scaled_by_local_minmax.csv')
    ids = range(0, len(df.index), 2)
    for i in ids:  # each participant have 2 rows (1-right hand, 2-left hand)
        right_data = df.iloc[i]
        name = right_data["Participant"]
        source = right_data["Source"]
        exercise = right_data["Exercise"]
        right_data = right_data[4:].dropna().to_numpy()
        right_data = filtfilt(y, x, right_data)
        left_data = df.iloc[i+1]
        left_data = left_data[4:].dropna().to_numpy()
        left_data = filtfilt(y, x, left_data)
        plt.figure()
        plt.title(name)
        plt.plot(right_data, label="right hand")
        plt.plot(left_data, label="left hand")
        plt.legend(loc='lower right')
        plt.xlabel("Frame")
        plt.ylabel("Angle Degree")
        plt.savefig('Plots/Signal/Scaled/'+exercise+'/'+source+'/'+name+'.png')


def add_angle(path):
    f = open('CSV/Raw Data/trynewnamma_'+'open_and_close_arms'+'.csv', 'w', newline='')
    writer = csv.writer(f)
    writer.writerow(['Participant','hand'])

    workbook_name = os.listdir(path)
    participants = []
    for wbn in workbook_name:
        workbook_path = path + "//" + wbn
        wb = openpyxl.load_workbook(filename=workbook_path)
        sheets_name = wb.sheetnames
        for i in sheets_name:
            if "open_and_close_arms" in i and 'v2' not in i:
                sheet = wb[i]
                if wbn in participants:  # participants that tried again the same exercise, in the same session
                    p = wbn + '2'
                else:
                    p = wbn
                participants.append(p)
                # Loop over columns in the sheet
                right_angles = [p, "right"]
                left_angles = [p, "left"]
                for column in sheet.iter_cols(min_col=2):
                    column_data = [cell.value for cell in column]
                    RIGHT_SHOULDER_x = column_data[6]
                    RIGHT_SHOULDER_y = column_data[7]
                    RIGHT_SHOULDER_z = column_data[8]
                    RIGHT_ELBOW_x = column_data[10]
                    RIGHT_ELBOW_y = column_data[11]
                    RIGHT_ELBOW_z = column_data[12]
                    RIGHT_HAND_x = column_data[14]
                    RIGHT_HAND_y = column_data[15]
                    RIGHT_HAND_z = column_data[16]
                    LEFT_SHOULDER_x = column_data[18]
                    LEFT_SHOULDER_y = column_data[19]
                    LEFT_SHOULDER_z = column_data[20]
                    LEFT_ELBOW_x = column_data[22]
                    LEFT_ELBOW_y = column_data[23]
                    LEFT_ELBOW_z = column_data[24]
                    LEFT_HAND_x = column_data[26]
                    LEFT_HAND_y = column_data[27]
                    LEFT_HAND_z = column_data[28]

                    right = calc_angle(LEFT_SHOULDER_x, LEFT_SHOULDER_y, LEFT_SHOULDER_z,
                                       RIGHT_SHOULDER_x, RIGHT_SHOULDER_y, RIGHT_SHOULDER_z,
                                       RIGHT_HAND_x, RIGHT_HAND_y, RIGHT_HAND_z)
                    left = calc_angle(RIGHT_SHOULDER_x, RIGHT_SHOULDER_y, RIGHT_SHOULDER_z,
                                      LEFT_SHOULDER_x, LEFT_SHOULDER_y, LEFT_SHOULDER_z,
                                      LEFT_HAND_x, LEFT_HAND_y, LEFT_HAND_z)
                    right_angles.append(right)
                    left_angles.append(left)
                writer.writerow(right_angles)
                writer.writerow(left_angles)
    f.close()


def add_angle2(p):
    f = open('CSV/Raw Data/trynewnamma_'+'open_arms_and_forward'+'.csv', 'w', newline='')
    writer = csv.writer(f)
    writer.writerow(['Participant','hand'])
    participants = []
    # iterate over all the files in directory 'parent'
    for path,dirs,files in os.walk(p):
        for file in files:
            workbook_path = path+ "//" + file
            wb = openpyxl.load_workbook(filename=workbook_path)
            sheets_name = wb.sheetnames
            for i in sheets_name:
                if "open_arms_and_forward" in i and 'v2' not in i:
                    sheet = wb[i]
                    if file in participants:  # participants that tried again the same exercise, in the same session
                        p = file + '2'
                    else:
                        p = file
                    participants.append(p)
                    # Loop over columns in the sheet
                    right_angles = [p, "right"]
                    left_angles = [p, "left"]
                    for column in sheet.iter_cols(min_col=2):
                        column_data = [cell.value for cell in column]
                        RIGHT_SHOULDER_x = column_data[6]
                        RIGHT_SHOULDER_y = column_data[7]
                        RIGHT_SHOULDER_z = column_data[8]
                        RIGHT_ELBOW_x = column_data[10]
                        RIGHT_ELBOW_y = column_data[11]
                        RIGHT_ELBOW_z = column_data[12]
                        RIGHT_HAND_x = column_data[14]
                        RIGHT_HAND_y = column_data[15]
                        RIGHT_HAND_z = column_data[16]
                        LEFT_SHOULDER_x = column_data[18]
                        LEFT_SHOULDER_y = column_data[19]
                        LEFT_SHOULDER_z = column_data[20]
                        LEFT_ELBOW_x = column_data[22]
                        LEFT_ELBOW_y = column_data[23]
                        LEFT_ELBOW_z = column_data[24]
                        LEFT_HAND_x = column_data[26]
                        LEFT_HAND_y = column_data[27]
                        LEFT_HAND_z = column_data[28]

                        right = calc_angle(LEFT_SHOULDER_x, LEFT_SHOULDER_y, LEFT_SHOULDER_z,
                                           RIGHT_SHOULDER_x, RIGHT_SHOULDER_y, RIGHT_SHOULDER_z,
                                           RIGHT_HAND_x, RIGHT_HAND_y, RIGHT_HAND_z)
                        left = calc_angle(RIGHT_SHOULDER_x, RIGHT_SHOULDER_y, RIGHT_SHOULDER_z,
                                          LEFT_SHOULDER_x, LEFT_SHOULDER_y, LEFT_SHOULDER_z,
                                          LEFT_HAND_x, LEFT_HAND_y, LEFT_HAND_z)
                        right_angles.append(right)
                        left_angles.append(left)
                    writer.writerow(right_angles)
                    writer.writerow(left_angles)
    f.close()


def calc_angle(joint1_x, joint1_y, joint1_z, joint2_x, joint2_y, joint2_z, joint3_x, joint3_y, joint3_z):
    joint1_x = float(joint1_x)
    joint1_y = float(joint1_y)
    joint1_z = float(joint1_z)/10
    joint2_x = float(joint2_x)
    joint2_y = float(joint2_y)
    joint2_z = float(joint2_z)/10
    joint3_x = float(joint3_x)
    joint3_y = float(joint3_y)
    joint3_z = float(joint3_z)/10

    a = np.array([joint1_x, joint1_y, joint1_z])  # First
    b = np.array([joint2_x, joint2_y, joint2_z])  # Mid
    c = np.array([joint3_x, joint3_y, joint3_z])  # End
    ba = a - b
    bc = c - b
    try:
        cosine_angle = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc))
        angle = np.arccos(cosine_angle)
        return round(np.degrees(angle), 2)

    except:
        print("could not calculate the angle")
        return None


def arrange_data(path, ex_name, row1, row2):
    # Read all excel data and transform to one table in csv file, including only angles
    f = open('CSV/Raw Data/val1'+ex_name+'.csv', 'w', newline='')
    writer = csv.writer(f)
    writer.writerow(['Participant','hand'])
    workbook_name = os.listdir(path)
    participants = []
    for wbn in workbook_name:
        workbook_path = path + "//" + wbn
        wb = openpyxl.load_workbook(filename=workbook_path)
        sheets_name = wb.sheetnames
        for i in sheets_name:
            if ex_name in i and 'v2' not in i:
                if ((ex_name == 'bend_elbows' and ('raise' not in i)) or ex_name != 'bend_elbows'):
                    ws_right = wb[i][row1]  # row 1 - right armpit angle in ex1
                    data_right = []
                    for cell in ws_right:
                        data_right.append(cell.value)
                    if wbn in participants:  # participants that tried again the same exercise, in the same session
                        p = wbn + '2'
                    else:
                        p = wbn
                    participants.append(p)
                    data_right = data_right[1:]
                    data_right = [p, "right"]+data_right
                    writer.writerow(data_right)
                    ws_left = wb[i][row2]
                    data_left = []
                    for cell in ws_left:
                        data_left.append(cell.value)
                    data_left = data_left[1:]
                    data_left = [p, "left"]+data_left
                    writer.writerow(data_left)
    f.close()


def scan_folder(p, data_source, ex_name, row1, row2):
    # Read all excel data and transform to one table in csv file, including only angles
    if ex_name == 'open_arms_bend_elbows':
        f = open('CSV\\Raw Data\\' + data_source + 'raise_arms_bend_elbows.csv', 'w', newline='')
    else:
        f = open('CSV\\Raw Data\\' + data_source + ex_name + '.csv', 'w', newline='')
    writer = csv.writer(f)
    writer.writerow(['Participant','hand'])
    # iterate over all the files in directory 'parent'
    for path,dirs,files in os.walk(p):
        for file in files:
            workbook_path = path+ "//" + file
            wb = openpyxl.load_workbook(filename=workbook_path)
            sheets_name = wb.sheetnames
            for i in sheets_name:
                if ex_name in i:
                    if ((ex_name == 'bend_elbows' and ('open' not in i)) or ex_name != 'bend_elbows'):
                        ws_right = wb[i][row1]  # row 1 - right armpit angle in ex1
                        data_right = []
                        for cell in ws_right:
                            data_right.append(cell.value)
                        data_right = data_right[1:]
                        data_right = [file, "right"]+data_right
                        writer.writerow(data_right)
                        ws_left = wb[i][row2]
                        data_left = []
                        for cell in ws_left:
                            data_left.append(cell.value)
                        data_left = data_left[1:]
                        data_left = [file, "left"]+data_left
                        writer.writerow(data_left)
    f.close()


def maya_exp():
    # 26 participants, 2 sessions for each
    path = u'C:\\Users\\mayak\\Documents\\לימודים\\תואר שני\\פרויקט גמר\\ניסויים COG\\דטה אקסל'
    exercises = [['raise_arms_horizontally', 26, 27], ['bend_elbows', 26, 27],['raise_arms_bend_elbows', 32, 33],
                 ['open_arms_and_forward', 34, 35]]
    for ex in exercises:
        arrange_data(path, ex[0], ex[1], ex[2])
        # create_graphs(ex_name[0], True, False)


def validation_exp():
    p = r'CSV\Raw Data\validationexp1'
    exercises = [['raise_arms_horizontally', 50, 51], ['bend_elbows', 26, 27],['raise_arms_bend_elbows', 50, 51],
                 ['open_and_close_arms', 52, 53],['open_and_close_arms_90',52,53],['raise_arms_forward',50,51]]
    for ex in exercises:
        arrange_data(p, ex[0], ex[1], ex[2])

    # OPEM_ARMS_AMD_CLOSE_90 contained in the file of OPEM_ARMS_AMD_CLOSE -
    # removing this so it will be sepreated
    OPEM_ARMS_AMD_CLOSE = pd.read_csv(r'CSV\Raw Data\val1open_and_close_arms.csv')
    df = OPEM_ARMS_AMD_CLOSE[~OPEM_ARMS_AMD_CLOSE['Participant'].str.endswith('2')]
    df.to_csv(r'CSV\Raw Data\val1open_and_close_arms.csv')


def naama_exp():
    # 16 older adults, 2 sessions for each
    p = r'C:\Users\mayak\Downloads\naamaexp'
    p = r'C:\Users\mayak\PycharmProjects\DataAnalysis\data\naamaexp'
    exercises = [['raise_arms_horizontally', 26, 27], ['bend_elbows', 26, 27],['open_arms_bend_elbows', 32, 33],
                 ['open_arms_and_forward', 34, 35]]
    for ex in exercises:
        scan_folder(p, 'naama_', ex[0], ex[1], ex[2])


def naama_pilot():
    p = r'C:\Users\mayak\Downloads\naamapilot'
    exercises = [['raise_arms_horizontally', 26, 27], ['bend_elbows', 26, 27],['open_arms_bend_elbows', 32, 33],
                 ['open_arms_and_forward', 34, 35]]
    for ex in exercises:
        scan_folder(p, 'naama_pilot_' , ex[0], ex[1], ex[2])


def extract_raw_data():
    maya_exp()
    naama_exp()
    naama_pilot()


if __name__ == '__main__':
    print('main')
    # extract_raw_data()
    # maya_exp()
    # naama_exp()
    # naama_pilot()

    p = r'C:\Users\mayak\PycharmProjects\DataAnalysis\data\naamaexp'
    add_angle2(p)

